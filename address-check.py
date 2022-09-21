from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import pandas as pd
import time
from openpyxl import load_workbook

# Get file name from user
file_loc = input("Enter excel file name: ") + ".xlsx"

# Chrome driver path
PATH = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(PATH)
delay = 3

# Site used to get address data
driver.get("https://neighborhoodinfo.lacity.org/")

# Store addresses from excel into a dataframe using pandas
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], usecols="A")

# Create a blank array(list) for district numbers
districtNumberList = []

# Iterate through addresses, search, and retrieve data
for i, row in df.iterrows():
        time.sleep(1)
        search = driver.find_element(By.ID, 'ita-acsf-neighborhoodinfo-modal-field-address')
        search.clear()
        search.send_keys(row)
        search.send_keys(Keys.RETURN)
        time.sleep(3)

        try:
            districtNumber = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div/div[3]/div[2]/div/main/section/div[1]/div/div/div[3]/table[1]/tr[5]/td[2]/a')))
            districtNumberList.append("CD " + ((districtNumber.text).split(" ")[1])[:-1])
            print(districtNumberList)
        except TimeoutException:
            districtNumberList.append("None")
        
# Can check result in terminal
print(districtNumberList)

# Post to Excel
wb = load_workbook(file_loc)
ws = wb.create_sheet("output")

for row in zip(districtNumberList):
    ws.append(row)

wb.save(file_loc)