from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import pandas as pd
import time
from openpyxl import load_workbook

#Chrome Driver Path
PATH = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(PATH)

#Site used to get address data
driver.get("https://www.google.com/maps/")

file_loc = "Pynoos2 - cleaned.xlsx"
df = pd.read_excel(file_loc, index_col=None, na_values=['NA'], usecols="A")

urlList = []

for i, row in df.iterrows():
         time.sleep(2)
         search = driver.find_element(By.ID, 'searchboxinput').clear()
         search = driver.find_element(By.ID, 'searchboxinput')
         time.sleep(1)
         search.click()
         time.sleep(1)
         search.send_keys(row)
         driver.find_element(By.ID, 'searchbox-searchbutton').click()
         time.sleep(2)
         currentURL = (driver.current_url)
         time.sleep(2)
         urlList.append(currentURL)
         print(urlList)

#Post to Excel
wb = load_workbook(file_loc)
ws = wb.create_sheet("googlemaps")

for row in zip(urlList):
   ws.append(row)

wb.save(file_loc)